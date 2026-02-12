"""
Módulo para leer y acceder a datos de archivos Excel de PLABACOM.
"""
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from pathlib import Path
from typing import Optional, Dict, List, Union
from datetime import datetime, date

from core.descargar_archivos import meses, meses_abrev


def encontrar_archivo_balance(anyo: int, mes: int, carpeta_base: str = "bd_data") -> Optional[Path]:
    """
    Encuentra el archivo Balance_XXYYD.xlsm basado en el año y mes.

    Args:
        anyo: Año del archivo
        mes: Mes del archivo (1-12)
        carpeta_base: Carpeta base donde buscar (por defecto "bd_data")

    Returns:
        Path del archivo encontrado, None si no existe
    """
    # Asegurar que anyo y mes sean enteros
    anyo = int(anyo)
    mes = int(mes)

    # Validar rango del mes
    if mes < 1 or mes > 12:
        print(f"[ERROR] Mes inválido: {mes}. Debe estar entre 1 y 12.")
        return None

    anyo_abrev = str(anyo)[-2:]
    mes_str = str(mes).zfill(2)

    # Construir nombre del archivo: Balance_2512D.xlsm (formato: Balance_YYMMD.xlsm)
    nombre_archivo = f"Balance_{anyo_abrev}{mes_str}D.xlsm"

    print(f"[INFO] Buscando archivo Balance para {mes}/{anyo}")
    print(f"  Nombre esperado: {nombre_archivo}")
    print(f"  Año abreviado: {anyo_abrev}, Mes: {mes_str}")

    # Buscar en la carpeta descomprimidos
    carpeta_descomprimidos = Path(carpeta_base) / "descomprimidos"

    if not carpeta_descomprimidos.exists():
        print(f"[ERROR] La carpeta no existe: {carpeta_descomprimidos.absolute()}")
        return None

    # Buscar en todas las carpetas que coincidan con el patrón
    # El patrón es: "01 Resultados_2512_BD01"
    patron_carpeta = f"*Resultados_{anyo_abrev}{mes_str}_BD01"
    print(f"  Patrón de carpeta: {patron_carpeta}")

    carpetas_encontradas = list(carpeta_descomprimidos.glob(patron_carpeta))
    print(f"  Carpetas encontradas con el patrón: {len(carpetas_encontradas)}")

    for carpeta in carpetas_encontradas:
        archivo_balance = carpeta / nombre_archivo
        if archivo_balance.exists():
            print(f"[OK] Archivo Balance encontrado: {archivo_balance}")
            print(f"  Ruta completa: {archivo_balance.absolute()}")
            return archivo_balance
        else:
            # Buscar también en subcarpetas (por si hay estructura anidada)
            for subcarpeta in carpeta.rglob("*"):
                if subcarpeta.is_dir():
                    archivo_balance = subcarpeta / nombre_archivo
                    if archivo_balance.exists():
                        print(f"[OK] Archivo Balance encontrado en subcarpeta: {archivo_balance}")
                        print(f"  Ruta completa: {archivo_balance.absolute()}")
                        return archivo_balance

    # Si no se encuentra, buscar en cualquier subcarpeta (búsqueda más amplia)
    print(f"  Realizando búsqueda amplia en todas las carpetas...")
    for carpeta in carpeta_descomprimidos.iterdir():
        if carpeta.is_dir():
            # Buscar directamente en la carpeta
            archivo_balance = carpeta / nombre_archivo
            if archivo_balance.exists():
                print(f"[OK] Archivo Balance encontrado: {archivo_balance}")
                print(f"  Ruta completa: {archivo_balance.absolute()}")
                return archivo_balance

            # Buscar en subcarpetas
            for archivo in carpeta.rglob(nombre_archivo):
                if archivo.is_file():
                    print(f"[OK] Archivo Balance encontrado: {archivo}")
                    print(f"  Ruta completa: {archivo.absolute()}")
                    return archivo

    # Listar archivos Balance disponibles para ayudar al usuario
    print(f"[ERROR] No se encontró el archivo Balance: {nombre_archivo}")
    print(f"  Buscado en: {carpeta_descomprimidos.absolute()}")

    # Mostrar archivos Balance disponibles
    archivos_balance = list(carpeta_descomprimidos.rglob("Balance_*.xlsm"))
    if archivos_balance:
        print(f"  Archivos Balance disponibles:")
        for archivo in archivos_balance[:5]:  # Mostrar máximo 5
            print(f"    - {archivo.name} (en {archivo.parent.name})")

    return None


# Mes abreviado para nombre del Anexo Potencia (Ene, Feb, ... Dic)
MESES_ANEXO_POTENCIA = {
    1: "Ene",
    2: "Feb",
    3: "Mar",
    4: "Abr",
    5: "May",
    6: "Jun",
    7: "Jul",
    8: "Ago",
    9: "Sep",
    10: "Oct",
    11: "Nov",
    12: "Dic",
}


def encontrar_archivo_anexo_potencia(
    anyo: int,
    mes: int,
    carpeta_base: str = "bd_data",
) -> Optional[Path]:
    """
    Encuentra el archivo Anexo 02.b Cuadros de Pago_Potencia_SEN_{Mes}{Year}_Simplificado
    en la carpeta descomprimida del ZIP de Potencia.

    Args:
        anyo: Año del archivo
        mes: Mes del archivo (1-12)
        carpeta_base: Carpeta base (por defecto "bd_data")

    Returns:
        Path del archivo encontrado, None si no existe
    """
    anyo = int(anyo)
    mes = int(mes)
    if mes < 1 or mes > 12:
        return None

    mes_anexo = MESES_ANEXO_POTENCIA.get(mes, "Dic")
    year2 = str(anyo)[-2:]
    # Buscar .xlsb y .xlsx
    nombres_posibles = [
        f"Anexo 02.b Cuadros de Pago_Potencia_SEN_{mes_anexo}{year2}_Simplificado.xlsb",
        f"Anexo 02.b Cuadros de Pago_Potencia_SEN_{mes_anexo}{year2}_Simplificado.xlsx",
    ]

    carpeta_descomprimidos = Path(carpeta_base) / "descomprimidos"
    if not carpeta_descomprimidos.exists():
        return None

    # Buscar en carpetas que contengan "Potencia"
    for carpeta in carpeta_descomprimidos.iterdir():
        if carpeta.is_dir() and "Potencia" in carpeta.name:
            for nombre in nombres_posibles:
                archivo = carpeta / nombre
                if archivo.exists():
                    return archivo
            # Buscar también en subcarpetas
            for archivo in carpeta.rglob("Anexo 02.b*Potencia*Simplificado*"):
                if archivo.suffix.lower() in (".xlsb", ".xlsx"):
                    return archivo

    return None


def encontrar_archivo_cuadros_pago_sscc(
    anyo: int,
    mes: int,
    carpeta_base: str = "bd_data",
) -> Optional[Path]:
    """
    Encuentra EXCEL 1_CUADROS_PAGO_SSCC_{YYMM}_def.xlsx en carpetas descomprimidas.
    Busca en carpetas SSCC y en toda la carpeta descomprimidos.

    Returns:
        Path del archivo, None si no existe
    """
    anyo = int(anyo)
    mes = int(mes)
    if mes < 1 or mes > 12:
        return None

    yymm = f"{str(anyo)[-2:]}{str(mes).zfill(2)}"
    # Patrones de nombre (PLABACOM puede usar distintas variantes)
    mes_abrev = meses_abrev.get(mes, "??")
    anyo_2 = str(anyo)[-2:]
    nombres = [
        f"1_CUADROS_PAGO_SSCC_{yymm}_def.xlsx",
        f"1_CUADROS_PAGO_SSCC_{yymm}_def.xlsb",
        f"1_CUADROS_PAGO_SSCC_{yymm}_def.xlsm",
        f"EXCEL 1_CUADROS_PAGO_SSCC_{yymm}_def.xlsx",
        f"EXCEL 1_CUADROS_PAGO_SSCC_{yymm}_def.xlsb",
        f"EXCEL 1_CUADROS_PAGO_SSCC_{yymm}_def.xlsm",
        f"1_CUADROS_PAGO_SSCC_{anyo_2}{mes_abrev}_def.xlsx",
        f"1_CUADROS_PAGO_SSCC_{anyo_2}{mes_abrev}_def.xlsb",
        f"1_CUADROS_PAGO_SSCC_{anyo_2}{mes_abrev}_def.xlsm",
        f"Cuadros de Pago_SSCC_{mes_abrev}{anyo_2}_def.xlsx",
        f"Cuadros de Pago_SSCC_{mes_abrev}{anyo_2}_def.xlsb",
        f"Cuadros de Pago_SSCC_{yymm}.xlsx",
        f"Cuadros de Pago_SSCC_{yymm}.xlsb",
    ]

    carpeta_base_path = Path(carpeta_base)
    carpeta_descomprimidos = carpeta_base_path / "descomprimidos"

    # 1) Buscar en bd_data/sscc (carpeta manual) y en bd_data raíz
    for carpeta_extra in [carpeta_base_path / "sscc", carpeta_base_path]:
        if carpeta_extra.exists():
            for nombre in nombres:
                archivo = carpeta_extra / nombre
                if archivo.exists():
                    return archivo
            for archivo in carpeta_extra.rglob("*"):
                if archivo.is_file() and archivo.suffix.lower() in (".xlsx", ".xlsb", ".xlsm"):
                    if "CUADROS_PAGO_SSCC" in archivo.name.upper() and yymm in archivo.name:
                        return archivo

    if not carpeta_descomprimidos.exists():
        return None

    # 2) Buscar por nombre exacto y por patrón en descomprimidos
    # Prioridad: PLABACOM_..._SSCC_Balance_SSCC_... (ej: PLABACOM_2025_12_Diciembre_SSCC_Balance_SSCC_2025_dic_def)
    for carpeta in carpeta_descomprimidos.iterdir():
        if carpeta.is_dir() and "SSCC" in carpeta.name:
            for nombre in nombres:
                archivo = carpeta / nombre
                if archivo.exists():
                    return archivo
            for archivo in carpeta.rglob("*"):
                if archivo.is_file() and archivo.suffix.lower() in (".xlsx", ".xlsb", ".xlsm"):
                    name_upper = archivo.name.upper()
                    name_lower = archivo.name.lower()
                    # Cuadros de Pago SSCC (variantes: 1_CUADROS_PAGO_SSCC, Cuadros de Pago_SSCC, etc.)
                    tiene_sscc_pago = (
                        ("PAGO" in name_upper and "SSCC" in name_upper)
                        or ("CUADROS" in name_upper and "SSCC" in name_upper)
                        or "cuadros_pago_sscc" in name_lower
                    )
                    # Período: 2512, dic25, dic, etc.
                    tiene_periodo = (
                        yymm in archivo.name
                        or f"{mes_abrev}{anyo_2}" in name_lower
                        or f"{anyo_2}{mes_abrev}" in name_lower
                    )
                    if tiene_sscc_pago and tiene_periodo:
                        return archivo

    # Búsqueda amplia
    for archivo in carpeta_descomprimidos.rglob("*"):
        if archivo.is_file() and archivo.suffix.lower() in (".xlsx", ".xlsb", ".xlsm"):
            if "CUADROS_PAGO_SSCC" in archivo.name.upper() and yymm in archivo.name:
                return archivo
    return None


def leer_total_ingresos_sscc(
    anyo: int,
    mes: int,
    nombre_empresa: str = "",
) -> Optional[float]:
    """
    Lee TOTAL INGRESOS POR SSCC CLP desde EXCEL 1_CUADROS_PAGO_SSCC, hoja CPI_.
    Filtra por Nemotecnico Deudor = nombre_empresa y suma columna Monto.
    Busca el archivo automáticamente en bd_data/descomprimidos.

    Returns:
        Valor en CLP, None si no se encuentra
    """
    archivo = encontrar_archivo_cuadros_pago_sscc(anyo, mes)
    if archivo is None:
        print(f"[WARNING] No se encontró EXCEL 1_CUADROS_PAGO_SSCC para {mes}/{anyo}")
        return None

    nombre_empresa_norm = (
        nombre_empresa.strip().upper().replace(" ", "_") if nombre_empresa else ""
    )
    if not nombre_empresa_norm:
        print("[WARNING] TOTAL INGRESOS POR SSCC: ingrese Empresa para filtrar por Nemotecnico Deudor")
        return None

    try:
        # CPI_ tiene encabezados en fila 5 (antes hay metadata: Coordinador, Concepto, etc.)
        kw = {"sheet_name": "CPI_", "header": None}
        if archivo.suffix.lower() == ".xlsb":
            kw["engine"] = "pyxlsb"
        df_raw = pd.read_excel(archivo, **kw)
    except Exception as e:
        print(f"[WARNING] Error leyendo CPI_: {e}")
        return None

    # Buscar fila de encabezados (contiene "Nemotecnico Deudor" y "Monto")
    fila_header = None
    for i in range(min(15, len(df_raw))):
        fila_str = " ".join(str(v) for v in df_raw.iloc[i].values if pd.notna(v)).lower()
        fila_str = fila_str.replace("ó", "o").replace("í", "i")
        if "nemotecnico" in fila_str and "deudor" in fila_str and "monto" in fila_str:
            fila_header = i
            break

    if fila_header is None:
        print(f"[WARNING] No se encontró fila de encabezados (Nemotecnico Deudor, Monto) en CPI_")
        return None

    df = df_raw.iloc[fila_header + 1 :].copy()
    headers = df_raw.iloc[fila_header].values

    # Buscar índice de columna (por posición); col 12 repite headers concatenados → evita duplicados
    idx_deudor = None
    idx_monto = None
    for idx, val in enumerate(headers):
        if val is None or (isinstance(val, float) and pd.isna(val)):
            continue
        if idx >= 10:  # Columnas 0-6 son datos; 7+ suele ser metadata o duplicado
            break
        c_lower = str(val).lower().replace("ó", "o").replace("í", "i")
        if "nemotecnico" in c_lower and "deudor" in c_lower and idx_deudor is None:
            idx_deudor = idx
        elif "monto" in c_lower and "retencion" not in c_lower and idx_monto is None:
            idx_monto = idx

    if idx_deudor is None or idx_monto is None:
        print(f"[WARNING] No se encontraron columnas Nemotecnico Deudor o Monto en CPI_")
        return None

    # Usar iloc para evitar duplicados de nombres
    def _norm_empresa(s):
        return str(s).strip().upper().replace(" ", "_")

    col_deudor_vals = df.iloc[:, idx_deudor].apply(lambda x: _norm_empresa(x))
    mask = col_deudor_vals == nombre_empresa_norm
    df_filtrado = df.loc[mask]
    total = df_filtrado.iloc[:, idx_monto].apply(
        lambda v: _parsear_valor_monetario(v) or 0
    ).sum()

    print(
        f"[INFO] Leyendo TOTAL INGRESOS POR SSCC CLP desde: {archivo.name} (hoja CPI_, Nemotecnico Deudor)"
    )
    print(f"  -> Dato obtenido ({nombre_empresa}, Monto): {total:,.2f}")
    return float(total)


def leer_compra_venta_energia_gm_holdings(
    anyo: int,
    mes: int,
    nombre_empresa: str = "",
    nombre_barra: str = "",
) -> Optional[float]:
    """
    Lee Compra Venta Energia GM Holdings CLP desde Balance_25XXD, hoja Contratos,
    columna Total CLP. Opcionalmente filtra por empresa/barra si existen esas columnas.

    Returns:
        Valor en CLP, None si no se encuentra
    """
    archivo = encontrar_archivo_balance(anyo, mes)
    if archivo is None:
        print(
            f"[WARNING] No se encontró Balance para leer Compra Venta GM Holdings {mes}/{anyo}"
        )
        return None

    try:
        df = pd.read_excel(archivo, sheet_name="Contratos", header=0)
    except Exception as e:
        print(f"[WARNING] Error leyendo hoja Contratos: {e}")
        return None

    # Patrones de columna para Total/Venta CLP (ordenados por preferencia)
    _patrones_clp = [
        lambda x: "total" in x and "clp" in x,
        lambda x: "venta" in x and "clp" in x,
        lambda x: "venta[clp]" in x or "venta (clp)" in x,
        lambda x: "monto" in x and "clp" in x,
        lambda x: x.strip().lower() == "total clp",
        lambda x: x.strip().lower() == "clp",
    ]
    col_total = None
    for c in df.columns:
        c_lower = str(c).strip().lower()
        for pred in _patrones_clp:
            if pred(c_lower):
                col_total = c
                break
        if col_total is not None:
            break
    if col_total is None:
        print(f"[WARNING] No se encontró columna Total CLP en Contratos")
        print(f"  Columnas disponibles: {list(df.columns)}")
        return None

    df_guardar = df.copy()
    if nombre_empresa:
        for c in df.columns:
            c_lower = str(c).lower()
            if "empresa" in c_lower or "nombre_corto" in c_lower or "nemotecnico" in c_lower:
                df_guardar = df_guardar[
                    df_guardar[c].astype(str).str.strip().str.upper()
                    == nombre_empresa.strip().upper()
                ]
                break
    if nombre_barra:
        for c in df.columns:
            if "barra" in str(c).lower():
                df_guardar = df_guardar[
                    df_guardar[c].astype(str).str.strip().str.upper()
                    == nombre_barra.strip().upper()
                ]
                break

    total = df_guardar[col_total].apply(
        lambda v: _parsear_valor_monetario(v) or 0
    ).sum()

    print(
        f"[INFO] Leyendo Compra Venta Energia GM Holdings CLP desde: {archivo.name} (hoja Contratos)"
    )
    print(f"  -> Dato obtenido (Total CLP): {total:,.2f}")
    return float(total)


def _encontrar_hoja_por_patron(
    ruta: Path,
    patrones: List[str],
    mes: int,
    anyo: int,
) -> Optional[str]:
    """
    Busca una hoja cuyo nombre contenga todos los patrones y el mes/año.
    Dinámico: funciona con cualquier formato (Hoja 02.IT..., 02.IT...Dic-25, etc.).

    Args:
        ruta: Archivo Excel
        patrones: Lista de cadenas que deben aparecer (ej: ["02.IT", "POTENCIA"])
        mes: Mes (1-12)
        anyo: Año

    Returns:
        Nombre de la hoja encontrada, None si no hay coincidencia
    """
    mes_anexo = MESES_ANEXO_POTENCIA.get(mes, "Dic")
    year2 = str(anyo)[-2:]
    variantes_mes_anyo = [
        f"{mes_anexo}-{year2}",
        f"{mes_anexo}{year2}",
        f"{mes_anexo.lower()}-{year2}",
        f"{mes_anexo.lower()}{year2}",
        year2,
    ]
    try:
        if ruta.suffix.lower() == ".xlsb":
            try:
                from pyxlsb import open_workbook
                with open_workbook(str(ruta)) as wb:
                    sheet_names = list(wb.sheets)
            except ImportError:
                return None
        else:
            xl = pd.ExcelFile(ruta)
            sheet_names = xl.sheet_names
        for nombre in sheet_names:
            n_lower = str(nombre).lower()
            if not all(p.lower() in n_lower for p in patrones):
                continue
            if not any(v.lower() in n_lower for v in variantes_mes_anyo):
                continue
            return nombre
    except Exception:
        pass
    return None


def _debug_mostrar_contenido_hoja(
    ruta: Path,
    nombre_hoja: str,
    max_filas: int = 30,
    max_cols: int = 10,
) -> None:
    """Muestra contenido de la hoja para depurar cuando no se encuentra un concepto."""
    try:
        try:
            df = pd.read_excel(ruta, sheet_name=nombre_hoja, header=None)
        except ValueError:
            xl = pd.ExcelFile(ruta)
            hoja = next((s for s in xl.sheet_names if nombre_hoja.lower() in str(s).lower()), None)
            if hoja is None:
                print(f"  Hojas disponibles: {xl.sheet_names}")
                return
            df = pd.read_excel(ruta, sheet_name=hoja, header=None)
        print(f"[DEBUG] Contenido de hoja (primeras {max_filas} filas):")
        for i, row in df.head(max_filas).iterrows():
            vals = [str(v)[:25] for v in row.iloc[:max_cols].tolist()]
            txt = " | ".join(v for v in vals if v and v != "nan")
            if txt.strip():
                print(f"  Fila {i}: {txt}")
        # Buscar celdas que contengan IT o INGRESOS
        celdas_relevantes = []
        for i, row in df.head(50).iterrows():
            for j, v in enumerate(row):
                if v and ("IT" in str(v).upper() or "INGRESOS" in str(v).upper()):
                    celdas_relevantes.append((i, j, str(v)[:50]))
        if celdas_relevantes:
            print(f"  Celdas con 'IT' o 'INGRESOS': {celdas_relevantes[:15]}")
    except Exception as e:
        print(f"[DEBUG] Error leyendo hoja para debug: {e}")
        try:
            xl = pd.ExcelFile(ruta)
            print(f"  Hojas disponibles: {xl.sheet_names}")
        except Exception:
            pass


def _leer_valor_por_empresa_y_columna(
    ruta: Path,
    nombre_hoja: str,
    nombre_empresa: str,
    col_valor: str,
) -> Optional[float]:
    """
    Lee valor buscando la fila por USUARIOS/empresa y la columna por nombre (ej: Total).
    Para hoja 02.IT POTENCIA: col A=USUARIOS, col Total=valor por empresa.
    Detecta dinámicamente la fila de encabezados (puede no ser la primera).
    """
    try:
        try:
            df_raw = pd.read_excel(ruta, sheet_name=nombre_hoja, header=None)
        except ValueError:
            xl = pd.ExcelFile(ruta)
            hoja = next((s for s in xl.sheet_names if nombre_hoja.lower() in str(s).lower()), None)
            if hoja is None:
                print(f"[DEBUG] Hoja no encontrada para IT: {nombre_hoja}")
                return None
            df_raw = pd.read_excel(ruta, sheet_name=hoja, header=None)

        # Buscar fila donde la PRIMERA celda sea "USUARIOS" o "EMPRESA" (evitar "Nota: Usuarios Pagan")
        header_row = None
        primeras_validas = ("USUARIOS", "EMPRESA")
        for i in range(min(20, len(df_raw))):
            first_cell = df_raw.iloc[i].iloc[0] if len(df_raw.columns) > 0 else None
            if first_cell is not None and not pd.isna(first_cell):
                first_str = str(first_cell).strip().upper()
                if first_str in primeras_validas or first_str.startswith("USUARIOS"):
                    header_row = i
                    break

        if header_row is None:
            for i in range(min(20, len(df_raw))):
                for j in range(min(3, len(df_raw.columns))):
                    cell = df_raw.iloc[i].iloc[j] if j < len(df_raw.iloc[i]) else None
                    if cell is not None and str(cell).strip().upper() in primeras_validas:
                        header_row = i
                        break
                if header_row is not None:
                    break

        if header_row is None:
            print("[DEBUG] No se encontró fila de encabezados con USUARIOS/Empresa")
            return None

        df = df_raw.iloc[header_row:].copy()
        df.columns = [str(c).strip() if c is not None and not (isinstance(c, float) and pd.isna(c)) else "" for c in df.iloc[0].tolist()]
        df = df.iloc[1:].reset_index(drop=True)

        emp_norm = str(nombre_empresa).strip().replace(" ", "_").upper()
        if not emp_norm:
            return None

        # Filtrar por columna Usuarios (prioridad: "Usuarios" exacto, luego usuario/empresa/nombre)
        col_empresa = None
        for c in df.columns:
            c_lower = str(c).strip().lower()
            if c_lower == "usuarios":
                col_empresa = c
                break
        if col_empresa is None:
            for c in df.columns:
                if "usuario" in str(c).lower() or "empresa" in str(c).lower():
                    col_empresa = c
                    break
        if col_empresa is None:
            col_empresa = df.columns[0]

        col_target = None
        for idx, c in enumerate(df.columns):
            c_str = str(c).strip().lower() if c else ""
            if c_str and col_valor.lower() in c_str:
                col_target = df.columns[idx]
                break
        if col_target is None:
            # Total suele ser la última columna en estas tablas
            for idx in range(len(df.columns) - 1, -1, -1):
                c = df.columns[idx]
                if str(c).strip().lower() == "total":
                    col_target = c
                    break
        if col_target is None:
            print(f"[DEBUG] No se encontró columna '{col_valor}'. Primeras: {list(df.columns)[:5]}... últimas: {list(df.columns)[-3:]}")
            return None

        for idx, row in df.iterrows():
            celda = row.get(col_empresa)
            if isinstance(celda, pd.Series):
                celda = celda.dropna().iloc[0] if len(celda.dropna()) > 0 else None
            if celda is None or (isinstance(celda, float) and pd.isna(celda)):
                continue
            celda_norm = str(celda).strip().replace(" ", "_").upper()
            if (
                emp_norm in celda_norm
                or celda_norm in emp_norm
                or emp_norm.startswith(celda_norm)
                or celda_norm.startswith(emp_norm)
            ):
                val = row.get(col_target)
                if isinstance(val, pd.Series):
                    val = val.dropna().iloc[0] if len(val.dropna()) > 0 else None
                parsed = _parsear_valor_monetario(val)
                if parsed is not None:
                    return parsed

        print(f"[DEBUG] No se encontró fila para '{nombre_empresa}' en col '{col_empresa}'. "
              f"Columnas: {list(df.columns)[:8]}...")
        return None
    except Exception as e:
        print(f"[DEBUG] Error leyendo por empresa/columna: {e}")
        return None


def _leer_valor_por_columna(
    ruta: Path,
    texto_upper: str,
    nombre_hoja: Optional[str],
    excluir_upper: List[str],
    col_valor_lower: str,
) -> Optional[float]:
    """
    Lee valor buscando la fila por texto_concepto y la columna por nombre.
    Usado cuando el valor está en una columna específica (ej: "Total general").
    """
    try:
        if nombre_hoja:
            try:
                df = pd.read_excel(ruta, sheet_name=nombre_hoja, header=0)
            except ValueError:
                # Hoja no encontrada, buscar por coincidencia parcial
                xl = pd.ExcelFile(ruta)
                hoja_encontrada = None
                for s in xl.sheet_names:
                    if nombre_hoja.lower() in str(s).lower():
                        hoja_encontrada = s
                        break
                if hoja_encontrada is None:
                    return None
                df = pd.read_excel(ruta, sheet_name=hoja_encontrada, header=0)
        else:
            df = pd.read_excel(ruta, sheet_name=0, header=0)

        # Buscar columna que contenga el nombre (ej: "total general")
        col_target = None
        for c in df.columns:
            if col_valor_lower in str(c).strip().lower():
                col_target = c
                break
        if col_target is None:
            print(f"[WARNING] No se encontró columna '{col_valor_lower}' en la hoja")
            return None

        # Buscar fila con texto_concepto (excluyendo filas con excluir)
        for idx, row in df.iterrows():
            row_str = " ".join(str(v).upper() for v in row.dropna().astype(str))
            if texto_upper in row_str:
                if excluir_upper and any(ex in row_str for ex in excluir_upper):
                    continue
                val = row.get(col_target)
                parsed = _parsear_valor_monetario(val)
                if parsed is not None:
                    print(f"[INFO] Valor encontrado en columna '{col_target}': {parsed:,.2f}")
                    return parsed
        print(f"[WARNING] No se encontró fila con el concepto en columna '{col_target}'")
        return None
    except Exception as e:
        print(f"[WARNING] Error leyendo por columna: {e}")
        return None


def leer_valor_concepto_anexo_xlsb(
    ruta_anexo: Path,
    texto_concepto: str,
    nombre_hoja: Optional[str] = None,
    excluir_si_contiene: Optional[List[str]] = None,
    columna_valor: Optional[str] = None,
) -> Optional[float]:
    """
    Busca 'texto_concepto' en el Anexo y devuelve el valor numérico asociado.
    Por defecto, el valor está en la misma fila, columna adyacente a la derecha.
    Si columna_valor está definida, toma el valor de esa columna por nombre.

    Args:
        ruta_anexo: Ruta del archivo .xlsb o .xlsx
        texto_concepto: Texto a buscar (ej: "TOTAL INGRESOS POR POTENCIA FIRME CLP")
        nombre_hoja: Nombre de la hoja donde buscar (ej: "01.BALANCE POTENCIA Dic-25 def").
            Si None, busca en todas las hojas.
        excluir_si_contiene: Si la celda contiene alguna de estas cadenas, se omite.
        columna_valor: Nombre de la columna donde tomar el valor (ej: "Total general").
            Si None, usa la primera columna numérica a la derecha del concepto.

    Returns:
        Valor numérico encontrado, None si no se encuentra
    """
    ruta = Path(ruta_anexo)
    if not ruta.exists():
        return None

    texto_upper = texto_concepto.upper()
    excluir_upper = (
        [e.upper() for e in excluir_si_contiene] if excluir_si_contiene else []
    )
    col_valor_lower = columna_valor.strip().lower() if columna_valor else None

    # Si se especifica columna_valor, leer por encabezados (pandas o pyxlsb)
    if col_valor_lower:
        return _leer_valor_por_columna(
            ruta, texto_upper, nombre_hoja, excluir_upper, col_valor_lower
        )

    if ruta.suffix.lower() == ".xlsb":
        try:
            from pyxlsb import open_workbook
        except ImportError:
            print("[WARNING] pyxlsb no instalado. Ejecute: pip install pyxlsb")
            return None

        with open_workbook(str(ruta)) as wb:
            # Si se especifica hoja, buscar solo ahí; si no, recorrer todas
            if nombre_hoja:
                hojas_a_revisar = [nombre_hoja] if nombre_hoja in wb.sheets else []
            else:
                hojas_a_revisar = list(wb.sheets)
            if nombre_hoja and not hojas_a_revisar:
                # Buscar coincidencia aproximada (insensible a mayúsculas)
                for s in wb.sheets:
                    if nombre_hoja.lower() in str(s).lower():
                        hojas_a_revisar = [s]
                        break

            for sheet_name in hojas_a_revisar:
                if sheet_name not in wb.sheets:
                    continue
                with wb.get_sheet(sheet_name) as sheet:
                    for row in sheet.rows():
                        # Construir dict col -> valor por si las celdas son sparse
                        row_by_col = {}
                        for cell in row:
                            if cell is not None:
                                c = getattr(cell, "c", len(row_by_col))
                                row_by_col[c] = getattr(cell, "v", cell)

                        for col_idx, val in row_by_col.items():
                            valor_str = str(val).strip().upper()
                            if texto_upper in valor_str:
                                if excluir_si_contiene:
                                    if any(
                                        ex.upper() in valor_str
                                        for ex in excluir_si_contiene
                                    ):
                                        break  # omitir fila, siguiente fila
                                # Buscar valor numérico en columnas siguientes de la misma fila
                                cols_orden = sorted(row_by_col.keys())
                                for c in cols_orden:
                                    if c > col_idx:
                                        v = row_by_col[c]
                                        try:
                                            return float(v)
                                        except (TypeError, ValueError):
                                            pass
                                return None
        return None

    # Para .xlsx usar pandas/openpyxl
    try:
        if nombre_hoja:
            df_dict = {nombre_hoja: pd.read_excel(ruta, sheet_name=nombre_hoja, header=None)}
        else:
            df_dict = pd.read_excel(ruta, sheet_name=None, header=None)
    except Exception:
        return None

    excluir_upper = (
        [e.upper() for e in excluir_si_contiene] if excluir_si_contiene else []
    )

    for _sheet_name, df in df_dict.items():
        for _, row in df.iterrows():
            row_str = " ".join(str(v).upper() for v in row.dropna().astype(str))
            if texto_upper in row_str:
                if excluir_upper and any(ex in row_str for ex in excluir_upper):
                    continue
                for v in row:
                    try:
                        return float(v)
                    except (TypeError, ValueError):
                        continue
                # Buscar primer número en la fila
                for v in row:
                    if isinstance(v, (int, float)) and not pd.isna(v):
                        return float(v)
                    try:
                        return float(v)
                    except (TypeError, ValueError):
                        pass
    return None


def _parsear_valor_monetario(val) -> Optional[float]:
    """Convierte valor con formato 46.709.214 (punto como miles) a float."""
    if val is None:
        return None
    if isinstance(val, (int, float)) and not (isinstance(val, bool)):
        return float(val)
    s = str(val).strip().replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def leer_total_ingresos_potencia_firme_anexo(
    ruta_anexo: Path,
    nombre_hoja: str,
    nombre_empresa: str,
) -> Optional[float]:
    """
    Lee el valor TOTAL para una empresa desde la tabla Datos del Anexo Potencia.
    Estructura: col B=Empresa, col C=Potencia SEN, col D=TOTAL.

    Args:
        ruta_anexo: Ruta del archivo .xlsb
        nombre_hoja: Nombre de la hoja (ej: "01.BALANCE POTENCIA Dic-25 def")
        nombre_empresa: Nombre de la empresa (ej: VIENTOS_DE_RENAICO)

    Returns:
        Valor en CLP de la columna TOTAL, None si no se encuentra
    """
    ruta = Path(ruta_anexo)
    if not ruta.exists():
        return None

    nombre_empresa_upper = nombre_empresa.strip().upper()
    if not nombre_empresa_upper:
        return None

    if ruta.suffix.lower() == ".xlsb":
        try:
            from pyxlsb import open_workbook
        except ImportError:
            return None

        with open_workbook(str(ruta)) as wb:
            if nombre_hoja not in wb.sheets:
                for s in wb.sheets:
                    if nombre_hoja.lower() in str(s).lower():
                        nombre_hoja = s
                        break
                else:
                    return None

            with wb.get_sheet(nombre_hoja) as sheet:
                # Col B=Empresa (índice 1), Col D=TOTAL (índice 3)
                # pyxlsb usa columnas 0-based en Cell
                for row in sheet.rows():
                    row_by_col = {}
                    for cell in row:
                        if cell is not None:
                            c = getattr(cell, "c", -1)
                            row_by_col[c] = getattr(cell, "v", cell)

                    # Columna B (índice 1) = Empresa
                    emp_val = row_by_col.get(1) or row_by_col.get(0)
                    if emp_val is None:
                        continue
                    if str(emp_val).strip().upper() != nombre_empresa_upper:
                        continue

                    # Columna D (índice 3) = TOTAL; fallback C (índice 2) = Potencia SEN
                    total_val = row_by_col.get(3) or row_by_col.get(2)
                    if total_val is not None:
                        parsed = _parsear_valor_monetario(total_val)
                        if parsed is not None:
                            return parsed
                return None

    # Fallback xlsx con pandas
    try:
        df = pd.read_excel(ruta, sheet_name=nombre_hoja, header=None)
    except Exception:
        return None
    # Buscar columna Empresa (B=1) y TOTAL (D=3) - pandas usa 0-based
    for _, row in df.iterrows():
        emp_cell = row.iloc[1] if len(row) > 1 else None
        if emp_cell is None:
            continue
        if str(emp_cell).strip().upper() != nombre_empresa_upper:
            continue
        total_cell = row.iloc[3] if len(row) > 3 else row.iloc[2]
        return _parsear_valor_monetario(total_cell)
    return None


def leer_total_ingresos_potencia_firme(
    anyo: int,
    mes: int,
    nombre_empresa: str = "",
) -> Optional[float]:
    """
    Lee el valor TOTAL INGRESOS POR POTENCIA FIRME CLP desde el Anexo 02.b
    de la carpeta Potencia, hoja "01.BALANCE POTENCIA {Mes}-{Year} def".
    Busca la fila donde Empresa = nombre_empresa y devuelve el valor de la columna TOTAL.

    Args:
        anyo: Año
        mes: Mes (1-12)
        nombre_empresa: Nombre de la empresa (ej: VIENTOS_DE_RENAICO). Si vacío, usa fallback.

    Returns:
        Valor en CLP, None si no se encuentra
    """
    archivo = encontrar_archivo_anexo_potencia(anyo, mes)
    if archivo is None:
        print(f"[WARNING] No se encontró Anexo 02.b Potencia para {mes}/{anyo}")
        return None

    nombre_hoja = _encontrar_hoja_por_patron(
        archivo, patrones=["01.BALANCE", "POTENCIA"], mes=mes, anyo=anyo
    )
    if nombre_hoja is None:
        nombre_hoja = f"01.BALANCE POTENCIA {MESES_ANEXO_POTENCIA.get(mes, 'Dic')}-{str(anyo)[-2:]} def"

    if nombre_empresa:
        valor = leer_total_ingresos_potencia_firme_anexo(
            archivo, nombre_hoja, nombre_empresa
        )
        if valor is not None:
            print(
                f"[INFO] Leyendo TOTAL INGRESOS POR POTENCIA FIRME CLP desde: {archivo.name}"
            )
            print(
                f"  -> Dato obtenido ({nombre_empresa}, col TOTAL): {valor:,.2f}"
            )
        return valor

    # Fallback: buscar por texto "TOTAL INGRESOS POR POTENCIA FIRME CLP"
    print(f"[INFO] Leyendo TOTAL INGRESOS POR POTENCIA FIRME CLP desde: {archivo.name}")
    valor = leer_valor_concepto_anexo_xlsb(
        archivo,
        "TOTAL INGRESOS POR POTENCIA FIRME CLP",
        nombre_hoja=nombre_hoja,
    )
    if valor is not None:
        print(f"  -> Dato obtenido (TOTAL INGRESOS POR POTENCIA FIRME CLP): {valor:,.2f}")
    return valor


def leer_ingresos_por_it(
    anyo: int,
    mes: int,
    nombre_empresa: str = "",
) -> Optional[float]:
    """
    Lee INGRESOS POR IT POTENCIA desde Anexo 02.b, hoja 02.IT/ASIGNACIÓN IT POTENCIA.

    Estructura: col USUARIOS (empresa), col Total (valor por fila).
    Si nombre_empresa está definido, busca la fila donde USUARIOS = empresa
    y devuelve el valor de la columna Total.

    Returns:
        Valor en CLP, None si no se encuentra
    """
    archivo = encontrar_archivo_anexo_potencia(anyo, mes)
    if archivo is None:
        print(f"[WARNING] No se encontró Anexo 02.b Potencia para leer INGRESOS POR IT POTENCIA {mes}/{anyo}")
        return None

    # Buscar hoja dinámicamente (02.IT POTENCIA, 02.ASIGNACIÓN IT POTENCIA, etc.)
    nombre_hoja = _encontrar_hoja_por_patron(
        archivo, patrones=["02", "IT", "POTENCIA"], mes=mes, anyo=anyo
    )
    if nombre_hoja is None:
        try:
            xl = pd.ExcelFile(archivo)
            hojas = xl.sheet_names
            print(f"[WARNING] No se encontró hoja 02.IT POTENCIA para {mes}/{anyo}")
            print(f"  Hojas en archivo: {hojas[:10]}{'...' if len(hojas) > 10 else ''}")
        except Exception:
            print(f"[WARNING] No se encontró hoja 02.IT POTENCIA para {mes}/{anyo}")
        return None

    print(f"[INFO] Hoja IT POTENCIA encontrada: {nombre_hoja}")

    # Estructura: col USUARIOS (empresa), col Total (valor). Requiere nombre_empresa para filtrar.
    if not nombre_empresa.strip():
        print("[WARNING] INGRESOS POR IT POTENCIA: ingrese Empresa en la interfaz para filtrar por Usuarios")
    if nombre_empresa:
        valor = _leer_valor_por_empresa_y_columna(
            archivo,
            nombre_hoja,
            nombre_empresa.strip(),
            col_valor="Total",
        )
        if valor is not None:
            print(f"[INFO] Leyendo INGRESOS POR IT POTENCIA desde: {archivo.name} (hoja {nombre_hoja}, col Total)")
            print(f"  -> Dato obtenido ({nombre_empresa}): {valor:,.2f}")
            return valor

    # Fallback: buscar por texto en fila
    for texto in ["INGRESOS POR IT POTENCIA", "INGRESOS POR IT"]:
        valor = leer_valor_concepto_anexo_xlsb(
            archivo,
            texto,
            nombre_hoja=nombre_hoja,
            columna_valor="Total",
        )
        if valor is None:
            valor = leer_valor_concepto_anexo_xlsb(
                archivo,
                texto,
                nombre_hoja=nombre_hoja,
            )
        if valor is not None:
            print(f"[INFO] Leyendo INGRESOS POR IT POTENCIA desde: {archivo.name} (hoja {nombre_hoja})")
            print(f"  -> Dato obtenido: {valor:,.2f}")
            return valor

    print(f"[WARNING] No se encontró INGRESOS POR IT POTENCIA en hoja {nombre_hoja}")
    _debug_mostrar_contenido_hoja(archivo, nombre_hoja)
    return None


def leer_ingresos_por_potencia(
    anyo: int,
    mes: int,
    nombre_empresa: str = "",
) -> Optional[float]:
    """
    Lee INGRESOS POR POTENCIA desde Anexo 02.b, hoja 01.BALANCE POTENCIA {Mes}-{YY} def.

    Estructura: col Empresa, col TOTAL. Filtra por Empresa y devuelve el valor de TOTAL.

    Returns:
        Valor en CLP, None si no se encuentra
    """
    archivo = encontrar_archivo_anexo_potencia(anyo, mes)
    if archivo is None:
        print(f"[WARNING] No se encontró Anexo 02.b Potencia para leer INGRESOS POR POTENCIA {mes}/{anyo}")
        return None

    nombre_hoja = _encontrar_hoja_por_patron(
        archivo, patrones=["01.BALANCE", "POTENCIA"], mes=mes, anyo=anyo
    )
    if nombre_hoja is None:
        nombre_hoja = f"01.BALANCE POTENCIA {MESES_ANEXO_POTENCIA.get(mes, 'Dic')}-{str(anyo)[-2:]} def"

    # Filtrar por Empresa y columna TOTAL (misma estructura que TOTAL INGRESOS POTENCIA FIRME)
    if not nombre_empresa.strip():
        print("[WARNING] INGRESOS POR POTENCIA: ingrese Empresa en la interfaz para filtrar")
    if nombre_empresa.strip():
        valor = leer_total_ingresos_potencia_firme_anexo(
            archivo, nombre_hoja, nombre_empresa.strip()
        )
        if valor is not None:
            print(f"[INFO] Leyendo INGRESOS POR POTENCIA desde: {archivo.name} (hoja {nombre_hoja}, filtro Empresa)")
            print(f"  -> Dato obtenido ({nombre_empresa}): {valor:,.2f}")
            return valor

        valor = _leer_valor_por_empresa_y_columna(
            archivo, nombre_hoja, nombre_empresa.strip(), col_valor="TOTAL"
        )
        if valor is not None:
            print(f"[INFO] Leyendo INGRESOS POR POTENCIA desde: {archivo.name} (hoja {nombre_hoja})")
            print(f"  -> Dato obtenido ({nombre_empresa}): {valor:,.2f}")
            return valor

    # Fallback: buscar por texto "INGRESOS POR POTENCIA"
    valor = leer_valor_concepto_anexo_xlsb(
        archivo,
        "INGRESOS POR POTENCIA",
        nombre_hoja=nombre_hoja,
        excluir_si_contiene=["FIRME"],
        columna_valor="Total general",
    )
    if valor is None:
        valor = leer_valor_concepto_anexo_xlsb(
            archivo,
            "INGRESOS POR POTENCIA",
            nombre_hoja=nombre_hoja,
            excluir_si_contiene=["FIRME"],
            columna_valor="TOTAL",
        )
    if valor is not None:
        print(f"[INFO] Leyendo INGRESOS POR POTENCIA desde: {archivo.name} (hoja {nombre_hoja})")
        print(f"  -> Dato obtenido: {valor:,.2f}")
    return valor


def leer_excel_pandas(ruta_archivo: Union[str, Path], hoja: Optional[str] = None, header: Optional[int] = None) -> pd.DataFrame:
    """
    Lee un archivo Excel usando pandas.

    Args:
        ruta_archivo: Ruta del archivo Excel
        hoja: Nombre de la hoja a leer (si None, lee la primera)
        header: Fila a usar como encabezado (si None, usa la primera fila)

    Returns:
        DataFrame con los datos
    """
    ruta_archivo = Path(ruta_archivo)
    if not ruta_archivo.exists():
        raise FileNotFoundError(f"El archivo no existe: {ruta_archivo}")

    if hoja:
        df = pd.read_excel(ruta_archivo, sheet_name=hoja, header=header)
    else:
        df = pd.read_excel(ruta_archivo, header=header)

    return df


def obtener_hojas_excel(ruta_archivo: Union[str, Path]) -> List[str]:
    """
    Obtiene la lista de hojas disponibles en un archivo Excel.

    Args:
        ruta_archivo: Ruta del archivo Excel

    Returns:
        Lista de nombres de hojas
    """
    ruta_archivo = Path(ruta_archivo)
    if not ruta_archivo.exists():
        raise FileNotFoundError(f"El archivo no existe: {ruta_archivo}")

    # Usar openpyxl para obtener las hojas (mejor para .xlsm)
    wb = openpyxl.load_workbook(ruta_archivo, read_only=True, data_only=True)
    hojas = wb.sheetnames
    wb.close()

    return hojas


def leer_celda_excel(ruta_archivo: Union[str, Path], hoja: str, celda: str):
    """
    Lee el valor de una celda específica de un archivo Excel.

    Args:
        ruta_archivo: Ruta del archivo Excel
        hoja: Nombre de la hoja
        celda: Referencia de celda (ej: "A1", "B5")

    Returns:
        Valor de la celda
    """
    ruta_archivo = Path(ruta_archivo)
    if not ruta_archivo.exists():
        raise FileNotFoundError(f"El archivo no existe: {ruta_archivo}")

    wb = openpyxl.load_workbook(ruta_archivo, read_only=True, data_only=True)
    ws = wb[hoja]
    valor = ws[celda].value
    wb.close()

    return valor


def leer_rango_excel(ruta_archivo: Union[str, Path], hoja: str, rango: str) -> pd.DataFrame:
    """
    Lee un rango específico de celdas de un archivo Excel.

    Args:
        ruta_archivo: Ruta del archivo Excel
        hoja: Nombre de la hoja
        rango: Rango de celdas (ej: "A1:C10")

    Returns:
        DataFrame con los datos del rango
    """
    ruta_archivo = Path(ruta_archivo)
    if not ruta_archivo.exists():
        raise FileNotFoundError(f"El archivo no existe: {ruta_archivo}")

    wb = openpyxl.load_workbook(ruta_archivo, read_only=True, data_only=True)
    ws = wb[hoja]

    # Leer el rango
    datos = []
    for fila in ws[rango]:
        datos.append([celda.value for celda in fila])

    wb.close()

    # Convertir a DataFrame
    df = pd.DataFrame(datos)

    return df


class LectorBalance:
    """
    Clase para leer y acceder a datos del archivo Balance de PLABACOM.
    """

    def __init__(self, anyo: int, mes: int, carpeta_base: str = "bd_data"):
        """
        Inicializa el lector para un año y mes específicos.

        Args:
            anyo: Año del archivo
            mes: Mes del archivo (1-12)
            carpeta_base: Carpeta base donde buscar
        """
        self.anyo = anyo
        self.mes = mes
        self.carpeta_base = carpeta_base
        print(f"\nBuscando archivo Balance para {mes}/{anyo}...")
        self.ruta_archivo = encontrar_archivo_balance(anyo, mes, carpeta_base)

        if self.ruta_archivo is None:
            print(f"[ERROR] No se encontró el archivo Balance para {mes}/{anyo} en {carpeta_base}")
            raise FileNotFoundError(f"No se encontró el archivo Balance para {mes}/{anyo} en {carpeta_base}")

        print(f"[OK] LectorBalance inicializado correctamente")
        print(f"  Archivo: {self.ruta_archivo.name}")
        print(f"  Ubicación: {self.ruta_archivo.parent}")

        self._hojas = None
        self._wb = None

    def obtener_hojas(self) -> List[str]:
        """Obtiene la lista de hojas disponibles."""
        if self._hojas is None:
            self._hojas = obtener_hojas_excel(self.ruta_archivo)
        return self._hojas

    def leer_hoja(self, nombre_hoja: str, header: Optional[int] = None) -> pd.DataFrame:
        """
        Lee una hoja completa del archivo Excel.

        Args:
            nombre_hoja: Nombre de la hoja a leer
            header: Fila a usar como encabezado (si None, usa la primera fila)

        Returns:
            DataFrame con los datos de la hoja
        """
        return leer_excel_pandas(self.ruta_archivo, hoja=nombre_hoja, header=header)

    def leer_celda(self, hoja: str, celda: str):
        """
        Lee el valor de una celda específica.

        Args:
            hoja: Nombre de la hoja
            celda: Referencia de celda (ej: "A1", "B5")

        Returns:
            Valor de la celda
        """
        return leer_celda_excel(self.ruta_archivo, hoja, celda)

    def leer_rango(self, hoja: str, rango: str) -> pd.DataFrame:
        """
        Lee un rango específico de celdas.

        Args:
            hoja: Nombre de la hoja
            rango: Rango de celdas (ej: "A1:C10")

        Returns:
            DataFrame con los datos del rango
        """
        return leer_rango_excel(self.ruta_archivo, hoja, rango)

    def _detectar_fila_encabezados(self, nombre_hoja: str, max_filas: int = 20) -> Optional[int]:
        """
        Detecta automáticamente la fila que contiene los encabezados 'barra' y 'monetario'.

        Args:
            nombre_hoja: Nombre de la hoja a analizar
            max_filas: Número máximo de filas a revisar

        Returns:
            Número de fila que contiene los encabezados, None si no se encuentra
        """
        # Leer las primeras filas sin encabezados para buscar manualmente
        df_temp = leer_excel_pandas(self.ruta_archivo, hoja=nombre_hoja, header=None)

        # Buscar en las primeras filas
        for fila_idx in range(min(max_filas, len(df_temp))):
            fila = df_temp.iloc[fila_idx]
            # Convertir la fila a strings y buscar 'barra' y 'monetario' (insensible a mayúsculas)
            valores_fila = [str(val).lower() if val is not None else "" for val in fila.values]

            tiene_barra = any("barra" in str(val).lower() for val in valores_fila)
            tiene_monetario = any("monetario" in str(val).lower() for val in valores_fila)

            if tiene_barra and tiene_monetario:
                print(f"[OK] Encabezados detectados en la fila {fila_idx}")
                return fila_idx

        return None

    def leer_balance_valorizado(self, header: Optional[int] = None) -> pd.DataFrame:
        """
        Lee la hoja "Balance Valorizado" del archivo Balance.
        Si header es None, detecta automáticamente la fila de encabezados.

        Args:
            header: Fila a usar como encabezado (si None, detecta automáticamente)

        Returns:
            DataFrame con los datos de la hoja "Balance Valorizado"
        """
        nombre_hoja = "Balance Valorizado"
        print(f"\nLeyendo hoja: {nombre_hoja}")

        # Verificar que la hoja existe (búsqueda insensible a mayúsculas)
        hojas_disponibles = self.obtener_hojas()

        # Buscar la hoja (insensible a mayúsculas)
        hoja_encontrada = None
        for hoja in hojas_disponibles:
            if hoja.lower() == nombre_hoja.lower():
                hoja_encontrada = hoja
                break

        if hoja_encontrada is None:
            print(f"[ERROR] Advertencia: La hoja '{nombre_hoja}' no se encontró en el archivo")
            print(f"  Hojas disponibles: {', '.join(hojas_disponibles[:10])}...")
            raise ValueError(f"La hoja '{nombre_hoja}' no existe en el archivo")

        # Usar el nombre exacto de la hoja encontrada
        nombre_hoja = hoja_encontrada

        print(f"[OK] Hoja '{nombre_hoja}' encontrada")

        # Si no se especifica header, detectarlo automáticamente
        if header is None:
            print("  Detectando automáticamente la fila de encabezados...")
            header = self._detectar_fila_encabezados(nombre_hoja)
            if header is None:
                print("[WARNING] No se pudo detectar automáticamente la fila de encabezados, usando fila 0")
                header = 0
            else:
                print(f"  Usando fila {header} como encabezados")

        df = self.leer_hoja(nombre_hoja, header=header)
        print(f"  Filas leídas: {len(df)}")
        print(f"  Columnas: {len(df.columns)}")

        # Mostrar las primeras columnas para verificación
        print(f"  Primeras columnas: {', '.join([str(col) for col in df.columns[:5]])}...")

        # Verificar que tiene las columnas necesarias
        columnas_lower = [str(col).lower() for col in df.columns]
        if "barra" not in columnas_lower:
            print("[WARNING] No se encontró la columna 'barra' en los encabezados")
            print(f"  Columnas disponibles: {', '.join([str(col) for col in df.columns[:10]])}...")
        if "monetario" not in columnas_lower:
            print("[WARNING] No se encontró la columna 'monetario' en los encabezados")
            print(f"  Columnas disponibles: {', '.join([str(col) for col in df.columns[:10]])}...")

        return df

    def obtener_columna(self, df: pd.DataFrame, nombre_columna: str, mostrar_todos: bool = True) -> pd.Series:
        """
        Obtiene una columna específica del DataFrame.

        Args:
            df: DataFrame del cual extraer la columna
            nombre_columna: Nombre de la columna a obtener (búsqueda insensible a mayúsculas)
            mostrar_todos: Si mostrar todos los elementos (True) o solo un resumen (False)

        Returns:
            Serie con los valores de la columna
        """
        # Buscar la columna (insensible a mayúsculas)
        columna_encontrada = None
        for col in df.columns:
            if str(col).lower() == nombre_columna.lower():
                columna_encontrada = col
                break

        if columna_encontrada is None:
            print(f"[ERROR] La columna '{nombre_columna}' no se encontró")
            print(f"  Columnas disponibles: {', '.join([str(c) for c in df.columns[:10]])}...")
            raise ValueError(f"La columna '{nombre_columna}' no existe en el DataFrame")

        print(f"\n[OK] Columna '{columna_encontrada}' encontrada")
        serie = df[columna_encontrada]

        if mostrar_todos:
            print(f"\nTodos los elementos de la columna '{columna_encontrada}':")
            print("=" * 60)
            for idx, valor in enumerate(serie, start=1):
                print(f"{idx:4d}. {valor}")
            print("=" * 60)
            print(f"\nTotal de elementos: {len(serie)}")
            print(f"Valores no nulos: {serie.notna().sum()}")
            print(f"Valores nulos: {serie.isna().sum()}")
        else:
            print(f"\nResumen de la columna '{columna_encontrada}':")
            print(serie.describe())

        return serie

    def buscar_por_barra(self, df: pd.DataFrame, nombre_barra: str) -> pd.DataFrame:
        """
        Busca todos los registros que corresponden a una barra específica y muestra los valores monetarios.

        Args:
            df: DataFrame con los datos del Balance Valorizado
            nombre_barra: Nombre de la barra a buscar (búsqueda insensible a mayúsculas)

        Returns:
            DataFrame filtrado con los registros de esa barra
        """
        # Buscar la columna "barra" (insensible a mayúsculas)
        columna_barra = None
        for col in df.columns:
            if str(col).lower() == "barra":
                columna_barra = col
                break

        if columna_barra is None:
            print("[ERROR] La columna 'barra' no se encontró en el DataFrame")
            raise ValueError("La columna 'barra' no existe en el DataFrame")

        # Filtrar por el nombre de la barra (insensible a mayúsculas)
        df_filtrado = df[df[columna_barra].astype(str).str.lower() == nombre_barra.lower()]

        print(f"\n[OK] Búsqueda de barra: '{nombre_barra}'")
        print(f"  Registros encontrados: {len(df_filtrado)}")

        if len(df_filtrado) == 0:
            print(f"[ERROR] No se encontraron registros para la barra '{nombre_barra}'")
            # Mostrar algunas barras disponibles como sugerencia
            barras_unicas = df[columna_barra].dropna().unique()[:10]
            print(f"  Algunas barras disponibles: {', '.join([str(b) for b in barras_unicas])}...")
            return df_filtrado

        # Buscar la columna "monetario"
        columna_monetario = None
        for col in df.columns:
            if str(col).lower() == "monetario":
                columna_monetario = col
                break

        if columna_monetario:
            print(f"\nValores monetarios para la barra '{nombre_barra}':")
            print("=" * 60)
            valores_monetarios = df_filtrado[columna_monetario]
            for idx, (index_row, valor) in enumerate(valores_monetarios.items(), start=1):
                print(f"{idx:4d}. {valor}")
            print("=" * 60)

            # Estadísticas
            valores_no_nulos = valores_monetarios.dropna()
            if len(valores_no_nulos) > 0:
                suma_total = valores_no_nulos.sum()
                print("\nResumen:")
                print(f"  Total de registros: {len(valores_monetarios)}")
                print(f"  Valores no nulos: {len(valores_no_nulos)}")
                print(f"  Suma total: {suma_total:,.2f}")
                print(f"  Promedio: {valores_no_nulos.mean():,.2f}")
                print(f"  Mínimo: {valores_no_nulos.min():,.2f}")
                print(f"  Máximo: {valores_no_nulos.max():,.2f}")
        else:
            print("[ERROR] La columna 'monetario' no se encontró")

        return df_filtrado

    def guardar_en_plantilla(
        self,
        df_balance: pd.DataFrame,
        ruta_plantilla: str = "plantilla_base.xlsx",
        nombre_barra: Optional[str] = None,
        nombre_empresa: Optional[str] = None,
    ) -> bool:
        """
        Guarda los datos del Balance Valorizado en la plantilla Excel.
        Si se especifica una barra o empresa, guarda solo los datos filtrados (TODAS las columnas).
        Si no se especifica, agrupa por barra y suma los valores monetarios.

        Args:
            df_balance: DataFrame con los datos del Balance Valorizado
            ruta_plantilla: Ruta del archivo plantilla Excel
            nombre_barra: Nombre de la barra específica (opcional)
            nombre_empresa: Nombre de la empresa (busca en columna nombre_corto_empresa) (opcional)

        Returns:
            True si se guardó correctamente, False en caso contrario
        """
        try:
            ruta_plantilla = Path(ruta_plantilla)

            # Crear nombre de la hoja destino.
            # Para plantillas de cliente se usa siempre la hoja "Resultado".
            nombre_mes = meses[self.mes]
            nombre_hoja = "Resultado"

            print(f"\n[OK] Guardando datos en plantilla: {ruta_plantilla}")
            print(f"  Hoja: {nombre_hoja}")

            # Buscar columnas necesarias
            columna_barra = None
            columna_monetario = None
            columna_empresa = None

            for col in df_balance.columns:
                col_lower = str(col).lower()
                if col_lower == "barra":
                    columna_barra = col
                elif col_lower == "monetario":
                    columna_monetario = col
                elif col_lower == "nombre_corto_empresa" or col_lower == "nombre corto empresa":
                    columna_empresa = col

            if not columna_barra or not columna_monetario:
                print("[ERROR] No se encontraron las columnas 'barra' o 'monetario'")
                return False

            # Preparar datos
            if nombre_barra or nombre_empresa:
                # Filtrar por barra y/o empresa - guardar TODAS las columnas de las filas filtradas
                df_guardar = df_balance.copy()

                # Aplicar filtro de empresa si se especifica
                if nombre_empresa:
                    if columna_empresa is None:
                        print("[ERROR] No se encontró la columna 'nombre_corto_empresa'")
                        print(
                            f"  Columnas disponibles: {', '.join([str(c) for c in df_balance.columns[:20]])}..."
                        )
                        return False

                    df_guardar = df_guardar[
                        df_guardar[columna_empresa].astype(str).str.lower() == nombre_empresa.lower()
                    ]
                    print(f"  Filtrando por empresa: {nombre_empresa} (columna: {columna_empresa})")
                    print(f"  Filas después de filtrar por empresa: {len(df_guardar)}")

                # Aplicar filtro de barra si se especifica
                if nombre_barra:
                    df_guardar = df_guardar[
                        df_guardar[columna_barra].astype(str).str.lower() == nombre_barra.lower()
                    ]
                    print(f"  Filtrando por barra: {nombre_barra}")

                print(f"  Filas encontradas después de todos los filtros: {len(df_guardar)}")
                print(f"  Columnas a guardar: {len(df_guardar.columns)}")
                print(f"  Columnas: {', '.join([str(col) for col in df_guardar.columns[:15]])}...")

                if len(df_guardar) == 0:
                    print("[WARNING] No se encontraron registros con los filtros especificados")
                    print(f"  Empresa: {nombre_empresa if nombre_empresa else 'Todas'}")
                    print(f"  Barra: {nombre_barra if nombre_barra else 'Todas'}")
            else:
                # Agrupar por barra y sumar valores monetarios
                df_guardar = df_balance.groupby(columna_barra)[columna_monetario].sum().reset_index()
                df_guardar.columns = ["Barra", "Monetario"]
                print("  Agrupando por barra y sumando valores monetarios")

            # Cargar o crear el archivo Excel
            if ruta_plantilla.exists():
                wb = load_workbook(ruta_plantilla)
            else:
                wb = openpyxl.Workbook()
                # Eliminar la hoja por defecto si existe
                if "Sheet" in wb.sheetnames:
                    wb.remove(wb["Sheet"])

            # Si la plantilla ya tiene una hoja "Resultado", escribir solo los
            # totales en la celda correspondiente al mes, sin destruir el diseño.
            if "Resultado" in wb.sheetnames:
                ws_resultado = wb["Resultado"]
                try:
                    self._escribir_resumen_en_hoja_resultado(
                        ws_resultado,
                        df_guardar,
                        nombre_mes,
                        self.anyo,
                        columna_monetario,
                    )
                    wb.save(ruta_plantilla)
                    wb.close()
                    print("[OK] Datos de resumen escritos en hoja 'Resultado' existente")
                    return True
                except Exception as e:
                    print(
                        "[ERROR] No se pudo escribir en hoja 'Resultado' existente: "
                        f"{e}"
                    )
                    wb.close()
                    return False

            # Si no hay una hoja 'Resultado', usar el comportamiento estándar:
            # crear/actualizar la hoja y volcar la tabla completa.
            if nombre_hoja in wb.sheetnames:
                ws = wb[nombre_hoja]
                # Limpiar la hoja existente
                ws.delete_rows(1, ws.max_row)
                print(f"  Hoja '{nombre_hoja}' ya existe, actualizando...")
            else:
                ws = wb.create_sheet(nombre_hoja)
                print(f"  Creando nueva hoja '{nombre_hoja}'")

            # Escribir encabezados
            encabezados = list(df_guardar.columns)
            print(
                "  Guardando "
                f"{len(encabezados)} columnas: "
                f"{', '.join([str(col) for col in encabezados[:15]])}..."
            )
            ws.append(encabezados)

            # Formatear encabezados
            from openpyxl.styles import Font, PatternFill

            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font

            # Escribir datos - asegurar que se escriban todas las filas y columnas
            print(f"  Escribiendo {len(df_guardar)} filas con todas las columnas...")
            for idx, r in enumerate(dataframe_to_rows(df_guardar, index=False, header=False), 1):
                ws.append(r)
                if idx % 100 == 0:
                    print(f"    Procesadas {idx} filas...")

            # Ajustar ancho de columnas
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except Exception:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Guardar el archivo
            wb.save(ruta_plantilla)
            wb.close()

            print("[OK] Datos guardados exitosamente")
            print(f"  Total de filas escritas: {len(df_guardar)}")
            print(f"  Total de columnas escritas: {len(df_guardar.columns)}")
            if nombre_barra:
                print("  Todas las columnas de las filas filtradas fueron guardadas")
            print(f"  Archivo: {ruta_plantilla.absolute()}")

            # Verificar que se guardaron todas las columnas
            if (nombre_barra or nombre_empresa) and len(df_guardar.columns) < len(df_balance.columns):
                print(
                    f"[WARNING] Se guardaron {len(df_guardar.columns)} columnas, "
                    f"pero el DataFrame original tiene {len(df_balance.columns)}"
                )
            elif nombre_barra or nombre_empresa:
                print(f"[OK] Se guardaron todas las {len(df_guardar.columns)} columnas del DataFrame original")

            return True

        except Exception as e:
            print(f"[ERROR] Error al guardar en plantilla: {e}")
            import traceback

            traceback.print_exc()
            return False

    def _escribir_resumen_en_hoja_resultado(
        self,
        ws,
        df_guardar: pd.DataFrame,
        nombre_mes: str,
        anyo: int,
        columna_monetario,
    ) -> None:
        """
        Escribe el total monetario del DataFrame en la hoja 'Resultado' de una
        plantilla existente, en la intersección:
        - Fila del concepto "TOTAL INGRESOS POR POTENCIA FIRME CLP"
        - Columna del mes/año correspondiente (por ejemplo, 'ene-25').
        """
        # Calcular total monetario del DataFrame filtrado/agrupado
        total_monetario = (
            df_guardar[columna_monetario].dropna().astype(float).sum()
        )
        print(f"  Total monetario a escribir en plantilla: {total_monetario:,.2f}")

        # Construir encabezado de mes esperado (ej: 'ene-25')
        meses_abrev = {
            1: "ene",
            2: "feb",
            3: "mar",
            4: "abr",
            5: "may",
            6: "jun",
            7: "jul",
            8: "ago",
            9: "sep",
            10: "oct",
            11: "nov",
            12: "dic",
        }
        encabezado_mes = f"{meses_abrev[self.mes]}-{str(anyo)[-2:]}"
        print(f"  Buscando columna de mes con encabezado: '{encabezado_mes}'")

        # 1) Encontrar columna del mes en las primeras filas (típicamente fila de encabezados)
        col_mes_idx = None
        fila_encabezados_max = 15
        for fila in ws.iter_rows(min_row=1, max_row=fila_encabezados_max):
            for cell in fila:
                raw = cell.value
                # Encabezado como fecha real (p.ej. 01-12-2025)
                if isinstance(raw, (datetime, date)):
                    if raw.year == anyo and raw.month == self.mes:
                        col_mes_idx = cell.column
                        print(f"  Columna de mes (fecha) encontrada en {cell.coordinate}: {raw}")
                        break
                # Encabezado como texto
                else:
                    valor = str(raw).strip()
                    valor_norm = valor.lower()
                    # Coincidencia flexible: empieza con el texto del mes (por si hay sufijos como ' CLP')
                    if valor_norm.startswith(encabezado_mes.lower()):
                        col_mes_idx = cell.column
                        print(f"  Columna de mes (texto) encontrada en {cell.coordinate}: {valor}")
                        break
            if col_mes_idx is not None:
                break

        if col_mes_idx is None:
            # Si no existe la columna, crear una nueva al final de la fila de encabezados
            print(
                f"  No se encontró columna para el mes '{encabezado_mes}'. "
                "Se creará una nueva columna de mes."
            )
            # Buscar la primera fila de encabezados no vacía
            encabezado_row = None
            for fila in ws.iter_rows(min_row=1, max_row=fila_encabezados_max):
                # Considerar fila encabezado si tiene al menos una celda no vacía
                if any(c.value not in (None, "") for c in fila):
                    encabezado_row = fila[0].row
                    break

            if encabezado_row is None:
                raise ValueError(
                    "No se pudo determinar la fila de encabezados para crear la columna de mes"
                )

            # Nueva columna: después de la última columna con datos en esa fila
            last_col = ws.max_column + 1
            header_cell = ws.cell(row=encabezado_row, column=last_col)
            # Usar fecha real para el encabezado; el formato lo maneja la plantilla
            header_cell.value = datetime(anyo, self.mes, 1)
            col_mes_idx = last_col
            print(f"  Columna de mes creada en {header_cell.coordinate} con fecha {header_cell.value}")

        # 2) Encontrar fila del concepto "TOTAL INGRESOS POR POTENCIA FIRME CLP"
        texto_concepto = "TOTAL INGRESOS POR POTENCIA FIRME CLP"
        fila_concepto_idx = None
        fila_busqueda_max = ws.max_row

        for row in ws.iter_rows(min_row=1, max_row=fila_busqueda_max):
            for cell in row:
                valor = str(cell.value).strip().upper() if cell.value is not None else ""
                # Coincidencia flexible: que contenga el texto del concepto
                if texto_concepto in valor:
                    fila_concepto_idx = cell.row
                    print(f"  Fila de concepto encontrada en {cell.coordinate}")
                    break
            if fila_concepto_idx is not None:
                break

        if fila_concepto_idx is None:
            raise ValueError(
                f"No se encontró la fila con el concepto '{texto_concepto}' en la hoja Resultado"
            )

        # 3) Escribir el total monetario en la celda correspondiente
        celda_destino = ws.cell(row=fila_concepto_idx, column=col_mes_idx)
        print(f"  Escribiendo valor en celda {celda_destino.coordinate}")
        celda_destino.value = float(total_monetario)

    def __repr__(self):
        return f"LectorBalance(anyo={self.anyo}, mes={self.mes}, archivo={self.ruta_archivo.name})"


if __name__ == "__main__":
    # Ejemplo de uso básico
    try:
        lector = LectorBalance(2025, 12)
        print(f"Archivo encontrado: {lector.ruta_archivo}")
        print("\nHojas disponibles:")
        for hoja in lector.obtener_hojas():
            print(f"  - {hoja}")

        print("\n" + "=" * 50)
        df_balance = lector.leer_balance_valorizado()
        print("\nPrimeras filas del Balance Valorizado:")
        print(df_balance.head(10))
    except FileNotFoundError as e:
        print(f"Error: {e}")

