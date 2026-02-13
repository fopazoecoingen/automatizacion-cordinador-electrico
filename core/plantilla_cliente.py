"""
Módulo para escribir valores en la plantilla del cliente.
En Windows: usa win32com (Excel COM) para máxima fidelidad con la plantilla del cliente.
En Mac o sin Excel: fallback a openpyxl.
"""
import re
import shutil
import sys
import tempfile
import time
from copy import copy
from datetime import datetime, date
from pathlib import Path
from typing import List, Tuple, Union

# Constante Excel: pegar solo formatos (evita copiar fechas/valores indeseados)
XL_PASTE_FORMATS = -4122

# Variantes de nombres que pueden aparecer en plantillas de clientes.
# ATENCIÓN: INGRESOS POR POTENCIA, INGRESOS POR IT POTENCIA y TOTAL INGRESOS POR POTENCIA FIRME CLP
# son conceptos distintos; cada uno escribe en su propia fila.
VARIANTES_CONCEPTOS = {
    "INGRESOS POR IT POTENCIA": [
        "IT POTENCIA",
        "INGRESOS IT POTENCIA",
        "02. IT POTENCIA",
        "INGRESOS POR IT",
        "IT Potencia",
        "Ingresos por IT Potencia",
        "IT/POTENCIA",
        "IT-POTENCIA",
        "POR IT POTENCIA",
        "ASIGNACION IT POTENCIA",
    ],
    "INGRESOS POR POTENCIA": [
        "INGRESOS POTENCIA",
        "01. INGRESOS POR POTENCIA",
        "Ingresos por Potencia",
        "POR POTENCIA",
    ],
}
# Al buscar cada concepto, se excluyen celdas que contengan estos textos.
# Así no se confunde INGRESOS POR POTENCIA con TOTAL INGRESOS POR POTENCIA FIRME CLP (son distintos).
EXCLUIR_AL_BUSCAR = {
    "INGRESOS POR POTENCIA": ["FIRME"],
}


def _ruta_local_para_excel(ruta: Path) -> Tuple[Path, bool]:
    """
    Retorna (ruta_a_usar, usar_temp).
    Solo usa temp cuando la ruta da problemas con Excel COM (OneDrive, red, etc.).
    Trabajar directo en la ruta evita corrupción al copiar desde temp.
    """
    ruta_abs = ruta.resolve()
    ruta_str = str(ruta_abs).lower()
    # Ejecutable: siempre usar temp (Excel COM falla con rutas cuando corre desde .exe)
    es_ejecutable = getattr(sys, "frozen", False)
    # Rutas problemáticas: OneDrive, red, descargas
    ruta_problemática = (
        "onedrive" in ruta_str or
        "google drive" in ruta_str or
        ruta_str.startswith("\\\\") or
        "\\downloads\\" in ruta_str or
        "\\descargas\\" in ruta_str or
        ruta_str.endswith("\\downloads") or
        ruta_str.endswith("\\descargas")
    )
    problemáticas = es_ejecutable or ruta_problemática
    if problemáticas:
        import uuid
        temp_dir = Path(tempfile.gettempdir()) / "GeneradorInformeElectrico"
        temp_dir.mkdir(parents=True, exist_ok=True)
        temp_file = temp_dir / f"{ruta.stem}_{uuid.uuid4().hex[:8]}{ruta.suffix}"
        shutil.copy2(ruta_abs, temp_file)
        return temp_file, True
    return ruta_abs, False


def _escribir_con_win32(
    ruta: Path,
    anyo: int,
    mes: int,
    total_monetario: float,
    texto_concepto: str,
) -> None:
    """Escritura usando Excel vía COM (Windows). Respeta formato y estructura de la plantilla."""
    import win32com.client as win32

    # Trabajar en ruta local para evitar errores con OneDrive/Downloads y Excel COM
    ruta_trabajo, usar_temp = _ruta_local_para_excel(ruta)
    ruta_abrir = str(ruta_trabajo)

    excel = win32.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False  # Evitar diálogos de Excel que pueden bloquear

    try:
        wb = excel.Workbooks.Open(ruta_abrir)
        ws_resultado = None
        for sh in wb.Worksheets:
            if str(sh.Name).strip().lower() == "resultado":
                ws_resultado = sh
                break
        if ws_resultado is None:
            excel.Quit()
            raise RuntimeError("No se encontró la hoja 'Resultado' en la plantilla.")
        ws = ws_resultado

        used_range = ws.UsedRange
        # +20 margen: UsedRange puede no incluir la última fila; el concepto puede estar al borde
        max_row = max(used_range.Rows.Count, 50) + 20
        max_col = max(used_range.Columns.Count, 30)

        meses_abrev = {
            1: "ene", 2: "feb", 3: "mar", 4: "abr", 5: "may", 6: "jun",
            7: "jul", 8: "ago", 9: "sep", 10: "oct", 11: "nov", 12: "dic",
        }
        encabezado_mes_2 = f"{meses_abrev[mes]}-{str(anyo)[-2:]}"
        encabezado_mes_4 = f"{meses_abrev[mes]}-{anyo}"

        col_mes = None
        fila_encabezados_max = min(15, max_row)

        for r in range(1, fila_encabezados_max + 1):
            for c in range(1, max_col + 1):
                raw = ws.Cells(r, c).Value
                if raw is None:
                    continue
                if isinstance(raw, (datetime, date)):
                    if raw.year == anyo and raw.month == mes:
                        col_mes = c
                        break
                else:
                    valor = str(raw).strip().replace(" ", "").lower()
                    if valor.startswith(encabezado_mes_2.lower()) or valor.startswith(encabezado_mes_4.lower()):
                        col_mes = c
                        break
            if col_mes is not None:
                break

        if col_mes is None:
            encabezado_row = None
            for r in range(1, fila_encabezados_max + 1):
                if any(ws.Cells(r, c).Value is not None for c in range(1, max_col + 1)):
                    encabezado_row = r
                    break
            if encabezado_row is None:
                excel.Quit()
                raise RuntimeError("No pude determinar la fila de encabezados.")

            patron_mes = re.compile(
                r"^(ene|feb|mar|abr|may|jun|jul|ago|sep|oct|nov|dic)[\s\-]*\d{2,4}$",
                re.I,
            )
            columnas_mes = []
            for c in range(1, max_col + 1):
                raw = ws.Cells(encabezado_row, c).Value
                if raw is None:
                    continue
                if isinstance(raw, (datetime, date)):
                    columnas_mes.append(c)
                elif raw and patron_mes.match(str(raw).strip().replace(" ", "")):
                    columnas_mes.append(c)

            if columnas_mes:
                base_col = max(columnas_mes)
                new_col = base_col + 1
                ws.Columns(new_col).Insert(Shift=0)
                ws.Columns(base_col).Copy()
                ws.Columns(new_col).PasteSpecial(Paste=XL_PASTE_FORMATS)
                excel.CutCopyMode = False
                header_cell = ws.Cells(encabezado_row, new_col)
                header_cell.Value = datetime(anyo, mes, 1)
                try:
                    base_fmt = ws.Cells(encabezado_row, base_col).NumberFormat
                    if base_fmt and str(base_fmt).strip():
                        header_cell.NumberFormat = base_fmt
                except Exception:
                    pass
            else:
                col_donde_insertar = 2
                for r in range(1, min(max_row + 1, 100)):
                    for col_candidate in (1, 2):
                        raw = ws.Cells(r, col_candidate).Value
                        if raw and "TOTAL INGRESOS" in str(raw).upper():
                            col_donde_insertar = col_candidate + 1
                            break
                    else:
                        continue
                    break
                new_col = col_donde_insertar
                ws.Columns(new_col).Insert(Shift=0)
                header_cell = ws.Cells(encabezado_row, new_col)
                header_cell.Value = datetime(anyo, mes, 1)
                header_cell.NumberFormat = "mmm-yy"

            col_mes = new_col

        fila_concepto = None
        textos_a_buscar = [texto_concepto] + VARIANTES_CONCEPTOS.get(texto_concepto, [])
        excluir = EXCLUIR_AL_BUSCAR.get(texto_concepto, [])

        def _celda_valida(val: str) -> bool:
            val_upper = val.upper()
            if any(ex.upper() in val_upper for ex in excluir):
                return False
            return True

        for texto_buscar in textos_a_buscar:
            texto_upper = texto_buscar.upper()
            for col_concepto in (2, 1, 3, 4):  # B, A, C, D
                for r in range(1, max_row + 1):
                    raw = ws.Cells(r, col_concepto).Value
                    if raw is None:
                        continue
                    val = str(raw).strip().upper()
                    if texto_upper in val and _celda_valida(val):
                        fila_concepto = r
                        break
                if fila_concepto is not None:
                    break
            if fila_concepto is not None:
                break

        if fila_concepto is None:
            wb.Close(SaveChanges=False)
            excel.Quit()
            raise RuntimeError(
                f"No se encontró el campo '{texto_concepto}' en la hoja Resultado (columnas A-D). "
                "Revise el nombre exacto en la plantilla y añádalo en VARIANTES_CONCEPTOS en plantilla_cliente.py."
            )

        ws.Cells(fila_concepto, col_mes).Value = float(total_monetario)
        wb.Save()
        wb.Close(SaveChanges=True)

        # Si trabajamos en temp, copiar el resultado de vuelta al destino original
        if usar_temp:
            shutil.copy2(ruta_trabajo, ruta.resolve())
    finally:
        excel.Quit()
        excel.DisplayAlerts = True


def _escribir_con_openpyxl(
    ruta: Path,
    anyo: int,
    mes: int,
    total_monetario: float,
    texto_concepto: str,
) -> None:
    """Fallback con openpyxl (Mac o cuando no hay Excel/win32)."""
    from openpyxl import load_workbook
    from openpyxl.worksheet.worksheet import Worksheet

    MESES_ABREV = {
        1: "ene", 2: "feb", 3: "mar", 4: "abr", 5: "may", 6: "jun",
        7: "jul", 8: "ago", 9: "sep", 10: "oct", 11: "nov", 12: "dic",
    }
    _PATRON_MES = re.compile(r"^(ene|feb|mar|abr|may|jun|jul|ago|sep|oct|nov|dic)[\s\-]*\d{2,4}$", re.I)

    def _copiar_estilo(origen, destino):
        if hasattr(origen, "font") and origen.font:
            destino.font = copy(origen.font)
        if hasattr(origen, "border") and origen.border:
            destino.border = copy(origen.border)
        if hasattr(origen, "fill") and origen.fill:
            destino.fill = copy(origen.fill)
        if hasattr(origen, "number_format") and origen.number_format:
            destino.number_format = copy(origen.number_format)
        if hasattr(origen, "alignment") and origen.alignment:
            destino.alignment = copy(origen.alignment)

    def _es_mes(raw, ay, mo):
        if isinstance(raw, (datetime, date)):
            return raw.year == ay and raw.month == mo
        enc2 = f"{MESES_ABREV[mo]}-{str(ay)[-2:]}"
        enc4 = f"{MESES_ABREV[mo]}-{ay}"
        val = str(raw).strip().replace(" ", "").lower()
        return val.startswith(enc2.lower()) or val.startswith(enc4.lower())

    def _es_columna_mes(ws: Worksheet, fila: int, col: int) -> bool:
        raw = ws.cell(row=fila, column=col).value
        if raw is None:
            return False
        if isinstance(raw, (datetime, date)):
            return True
        return bool(_PATRON_MES.match(str(raw).strip().replace(" ", "")))

    wb = load_workbook(str(ruta), data_only=False)
    ws = None
    for name in wb.sheetnames:
        if str(name).strip().lower() == "resultado":
            ws = wb[name]
            break
    if ws is None:
        wb.close()
        raise RuntimeError("No se encontró la hoja 'Resultado' en la plantilla.")

    # +20 margen: conceptos al final de la hoja pueden quedar fuera si max_row no los incluye
    max_row = max(ws.max_row, 50) + 20
    max_col = max(ws.max_column, 30)
    fila_encabezados_max = min(15, max_row)

    col_mes = None
    encabezado_row = None
    for r in range(1, fila_encabezados_max + 1):
        for c in range(1, max_col + 1):
            raw = ws.cell(row=r, column=c).value
            if raw is None:
                continue
            if encabezado_row is None:
                encabezado_row = r
            if _es_mes(raw, anyo, mes):
                col_mes = c
                break
        if col_mes is not None:
            break
    if encabezado_row is None:
        encabezado_row = 1

    if col_mes is None:
        columnas_mes = [c for c in range(1, max_col + 1) if _es_columna_mes(ws, encabezado_row, c)]
        if columnas_mes:
            base_col = max(columnas_mes)
            new_col = base_col + 1
            ws.insert_cols(new_col)
            for row in range(1, max_row + 1):
                _copiar_estilo(ws.cell(row=row, column=base_col), ws.cell(row=row, column=new_col))
            h = ws.cell(row=encabezado_row, column=new_col)
            h.value = datetime(anyo, mes, 1)
            h.number_format = ws.cell(row=encabezado_row, column=base_col).number_format or "mmm-yy"
            col_mes = new_col
        else:
            col_ins = 2
            for r in range(1, min(max_row + 1, 100)):
                for cc in (1, 2):
                    raw = ws.cell(row=r, column=cc).value
                    if raw and "TOTAL INGRESOS" in str(raw).upper():
                        col_ins = cc + 1
                        break
                else:
                    continue
                break
            ws.insert_cols(col_ins)
            h = ws.cell(row=encabezado_row, column=col_ins)
            h.value = datetime(anyo, mes, 1)
            h.number_format = "mmm-yy"
            col_mes = col_ins

    fila_concepto = None
    textos_a_buscar = [texto_concepto] + VARIANTES_CONCEPTOS.get(texto_concepto, [])
    excluir = EXCLUIR_AL_BUSCAR.get(texto_concepto, [])

    def _celda_valida(val: str) -> bool:
        val_upper = val.upper()
        return not any(ex.upper() in val_upper for ex in excluir)

    for texto_buscar in textos_a_buscar:
        texto_upper = texto_buscar.upper()
        for col_concepto in (2, 1, 3, 4):  # B, A, C, D
            for r in range(1, max_row + 1):
                raw = ws.cell(row=r, column=col_concepto).value
                if raw:
                    val = str(raw).strip().upper()
                    if texto_upper in val and _celda_valida(val):
                        fila_concepto = r
                        break
            if fila_concepto is not None:
                break
        if fila_concepto is not None:
            break

    if fila_concepto is None:
        wb.close()
        raise RuntimeError(
            f"No se encontró el campo '{texto_concepto}' en la hoja Resultado (columnas A-D). "
            "Revise el nombre exacto en la plantilla y añádalo en VARIANTES_CONCEPTOS en plantilla_cliente.py."
        )

    ws.cell(row=fila_concepto, column=col_mes).value = float(total_monetario)
    wb.save(str(ruta))
    wb.close()


# Excel constantes (evita corrupción de drawing.xml al guardar)
XL_UPDATE_LINKS_NEVER = 0
XL_OPEN_XML_WORKBOOK = 51


def _escribir_todos_con_win32(
    ruta: Path,
    anyo: int,
    mes: int,
    pares_concepto_valor: List[Tuple[str, float]],
) -> None:
    """Escribe todos los conceptos en una sola sesión Excel (evita múltiples open/close que fallan en el exe)."""
    import win32com.client as win32

    # Inicializar COM (necesario cuando se ejecuta desde .exe empaquetado)
    try:
        import pythoncom
        pythoncom.CoInitialize()
    except ImportError:
        pass

    ruta_trabajo, usar_temp = _ruta_local_para_excel(ruta)
    ruta_abrir = str(ruta_trabajo.resolve())

    excel = win32.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    excel.ScreenUpdating = False

    try:
        # UpdateLinks=0 evita que Excel modifique enlaces; reduce corrupción de dibujos
        wb = excel.Workbooks.Open(ruta_abrir, UpdateLinks=XL_UPDATE_LINKS_NEVER)
        ws_resultado = None
        for sh in wb.Worksheets:
            if str(sh.Name).strip().lower() == "resultado":
                ws_resultado = sh
                break
        if ws_resultado is None:
            excel.Quit()
            raise RuntimeError("No se encontró la hoja 'Resultado' en la plantilla.")
        ws = ws_resultado

        used_range = ws.UsedRange
        max_row = max(used_range.Rows.Count, 50) + 20
        max_col = max(used_range.Columns.Count, 30)
        meses_abrev = {
            1: "ene", 2: "feb", 3: "mar", 4: "abr", 5: "may", 6: "jun",
            7: "jul", 8: "ago", 9: "sep", 10: "oct", 11: "nov", 12: "dic",
        }
        encabezado_mes_2 = f"{meses_abrev[mes]}-{str(anyo)[-2:]}"
        encabezado_mes_4 = f"{meses_abrev[mes]}-{anyo}"
        col_mes = None
        fila_encabezados_max = min(15, max_row)

        for r in range(1, fila_encabezados_max + 1):
            for c in range(1, max_col + 1):
                raw = ws.Cells(r, c).Value
                if raw is None:
                    continue
                if isinstance(raw, (datetime, date)):
                    if raw.year == anyo and raw.month == mes:
                        col_mes = c
                        break
                else:
                    valor = str(raw).strip().replace(" ", "").lower()
                    if valor.startswith(encabezado_mes_2.lower()) or valor.startswith(encabezado_mes_4.lower()):
                        col_mes = c
                        break
            if col_mes is not None:
                break

        if col_mes is None:
            encabezado_row = None
            for r in range(1, fila_encabezados_max + 1):
                if any(ws.Cells(r, c).Value is not None for c in range(1, max_col + 1)):
                    encabezado_row = r
                    break
            if encabezado_row is None:
                excel.Quit()
                raise RuntimeError("No pude determinar la fila de encabezados.")
            patron_mes = re.compile(
                r"^(ene|feb|mar|abr|may|jun|jul|ago|sep|oct|nov|dic)[\s\-]*\d{2,4}$",
                re.I,
            )
            columnas_mes = []
            for c in range(1, max_col + 1):
                raw = ws.Cells(encabezado_row, c).Value
                if raw is None:
                    continue
                if isinstance(raw, (datetime, date)):
                    columnas_mes.append(c)
                elif raw and patron_mes.match(str(raw).strip().replace(" ", "")):
                    columnas_mes.append(c)
            if columnas_mes:
                base_col = max(columnas_mes)
                new_col = base_col + 1
                ws.Columns(new_col).Insert(Shift=0)
                # Copiar solo rango usado (evita corrupción de drawing.xml)
                filas_copiar = min(max_row, used_range.Rows.Count + 30)
                rng_origen = ws.Range(ws.Cells(1, base_col), ws.Cells(filas_copiar, base_col))
                rng_dest = ws.Range(ws.Cells(1, new_col), ws.Cells(filas_copiar, new_col))
                rng_origen.Copy()
                rng_dest.PasteSpecial(Paste=XL_PASTE_FORMATS)
                excel.CutCopyMode = False
                header_cell = ws.Cells(encabezado_row, new_col)
                header_cell.Value = datetime(anyo, mes, 1)
                try:
                    base_fmt = ws.Cells(encabezado_row, base_col).NumberFormat
                    if base_fmt and str(base_fmt).strip():
                        header_cell.NumberFormat = base_fmt
                except Exception:
                    pass
            else:
                col_donde_insertar = 2
                for r in range(1, min(max_row + 1, 100)):
                    for col_candidate in (1, 2):
                        raw = ws.Cells(r, col_candidate).Value
                        if raw and "TOTAL INGRESOS" in str(raw).upper():
                            col_donde_insertar = col_candidate + 1
                            break
                    else:
                        continue
                    break
                new_col = col_donde_insertar
                ws.Columns(new_col).Insert(Shift=0)
                header_cell = ws.Cells(encabezado_row, new_col)
                header_cell.Value = datetime(anyo, mes, 1)
                header_cell.NumberFormat = "mmm-yy"
            col_mes = new_col

        for texto_concepto, total_monetario in pares_concepto_valor:
            fila_concepto = None
            textos_a_buscar = [texto_concepto] + VARIANTES_CONCEPTOS.get(texto_concepto, [])
            excluir = EXCLUIR_AL_BUSCAR.get(texto_concepto, [])

            def _celda_valida(val: str) -> bool:
                val_upper = val.upper()
                if any(ex.upper() in val_upper for ex in excluir):
                    return False
                return True

            for texto_buscar in textos_a_buscar:
                texto_upper = texto_buscar.upper()
                for col_concepto in (2, 1, 3, 4):
                    for r in range(1, max_row + 1):
                        raw = ws.Cells(r, col_concepto).Value
                        if raw is None:
                            continue
                        val = str(raw).strip().upper()
                        if texto_upper in val and _celda_valida(val):
                            fila_concepto = r
                            break
                    if fila_concepto is not None:
                        break
                if fila_concepto is not None:
                    break

            if fila_concepto is None:
                wb.Close(SaveChanges=False)
                excel.Quit()
                raise RuntimeError(
                    f"No se encontró el campo '{texto_concepto}' en la hoja Resultado (columnas A-D). "
                    "Revise el nombre exacto en la plantilla y añádalo en VARIANTES_CONCEPTOS en plantilla_cliente.py."
                )

            ws.Cells(fila_concepto, col_mes).Value = float(total_monetario)

        wb.Save()
        wb.Close(SaveChanges=True)
    finally:
        try:
            excel.ScreenUpdating = True
            excel.DisplayAlerts = True
            excel.Quit()
        except Exception:
            pass
        time.sleep(0.5)  # Dar tiempo a Excel para liberar el archivo
        if usar_temp:
            shutil.copy2(ruta_trabajo, ruta.resolve())


def escribir_total_en_resultado(
    ruta_archivo: Union[str, Path],
    anyo: int,
    mes: int,
    total_monetario: float,
    texto_concepto: str = "TOTAL INGRESOS POR POTENCIA FIRME CLP",
) -> None:
    """
    Escribe total_monetario en la hoja 'Resultado' de la plantilla del cliente.
    En Windows con Excel: usa win32com (máxima compatibilidad con la plantilla).
    En Mac o sin Excel: usa openpyxl como fallback.
    """
    ruta = Path(ruta_archivo)
    if not ruta.exists():
        raise FileNotFoundError(ruta)

    if sys.platform == "win32":
        try:
            _escribir_con_win32(ruta, anyo, mes, total_monetario, texto_concepto)
            return
        except ImportError:
            pass  # pywin32 no instalado, usar openpyxl

    _escribir_con_openpyxl(ruta, anyo, mes, total_monetario, texto_concepto)


def escribir_todos_en_resultado(
    ruta_archivo: Union[str, Path],
    anyo: int,
    mes: int,
    pares_concepto_valor: List[Tuple[str, float]],
) -> None:
    """
    Escribe todos los conceptos en una sola sesión de Excel (reduce errores COM en el ejecutable).
    pares_concepto_valor: lista de (texto_concepto, valor).
    """
    if not pares_concepto_valor:
        return
    ruta = Path(ruta_archivo)
    if not ruta.exists():
        raise FileNotFoundError(ruta)

    if sys.platform != "win32":
        for concepto, valor in pares_concepto_valor:
            _escribir_con_openpyxl(ruta, anyo, mes, valor, concepto)
        return

    try:
        _escribir_todos_con_win32(ruta, anyo, mes, pares_concepto_valor)
    except ImportError:
        for concepto, valor in pares_concepto_valor:
            _escribir_con_openpyxl(ruta, anyo, mes, valor, concepto)
